
from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import openpyxl


CLIENT_REGISTRY_HEADERS = [
    "client_id_optional",
    "client_name",
    "dp_owner",
    "next_call_date",
    "alignment_notes_doc_link",
    "kickoff_doc_link",
    "acct_info_doc_link",
    "client_dosssier_doc_link",
    "client_overview_link",
    "irg_doc_link",
    "initial_strategy_link",
    "current_quarterly_strategy_link",
    "basecamp_internal_ref",
    "basecamp_joint_ref",
    "ga4_target_ref",
    "gsc_target_ref",
    "primary_dashboard_ref_optional",
    "producer_notes_store_key_optional",
    "last_verified_date",
    "registry_notes_optional",
]

SOURCE_INDEX_HEADERS = [
    "client_id_or_name",
    "source_type",
    "source_class",
    "source_label",
    "source_ref",
    "sub_ref_or_tab_optional",
    "priority",
    "active",
    "notes_optional",
]

PRODUCER_NOTES_HEADERS = [
    "note_id",
    "client_id_or_name",
    "note_type",
    "title",
    "note_text",
    "workstream_optional",
    "related_source_ref_optional",
    "priority",
    "pinned",
    "starts_on_optional",
    "expires_on_optional",
    "author",
    "active",
    "notes_optional",
]

REQUIRED_SHEETS = [
    "Client Registry",
    "Source Index",
    "Producer Notes",
]

SUPPORTED_SOURCE_CLASSES = {
    "foundational_strategy",
    "active_strategy",
    "narrative_memory",
    "live_execution_truth",
    "human_nuance",
    "continuity",  # legacy value normalized later
}

YES_VALUES = {"yes", "y", "true", "1", "active"}


def _coerce_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return value


def _header_row(ws) -> List[str]:
    return [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def _rows_as_dicts(ws, headers: List[str]) -> List[Dict[str, Any]]:
    actual_headers = _header_row(ws)
    if actual_headers[:len(headers)] != headers:
        raise ValueError(
            f"Sheet '{ws.title}' headers do not match expected headers.\n"
            f"Expected: {headers}\nGot:      {actual_headers[:len(headers)]}"
        )
    results: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else None
            item[header] = _coerce_scalar(value)
        results.append(item)
    return results


def load_workbook_data(workbook_path: str | Path) -> Dict[str, List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path)
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

    clients = _rows_as_dicts(wb["Client Registry"], CLIENT_REGISTRY_HEADERS)
    sources = _rows_as_dicts(wb["Source Index"], SOURCE_INDEX_HEADERS)
    notes = _rows_as_dicts(wb["Producer Notes"], PRODUCER_NOTES_HEADERS)
    return {
        "clients": clients,
        "sources": sources,
        "producer_notes": notes,
    }


def _is_example_row(row: Dict[str, Any]) -> bool:
    return str(row.get("client_id_optional") or "").strip() == "optional_slug" or str(row.get("client_name") or "").strip().lower() == "example client"


def find_client(workbook_data: Dict[str, List[Dict[str, Any]]], client_selector: str) -> Dict[str, Any]:
    matches = []
    selector = client_selector.strip().lower()
    for row in workbook_data["clients"]:
        if _is_example_row(row):
            continue
        cid = str(row.get("client_id_optional") or "").strip().lower()
        cname = str(row.get("client_name") or "").strip().lower()
        if selector in {cid, cname}:
            matches.append(row)
    if not matches:
        raise ValueError(f"No client row matched selector: {client_selector}")
    if len(matches) > 1:
        raise ValueError(f"Multiple client rows matched selector: {client_selector}")
    return matches[0]


def _client_key_variants(client_row: Dict[str, Any]) -> set[str]:
    variants = set()
    for key in (client_row.get("client_id_optional"), client_row.get("client_name")):
        if key:
            variants.add(str(key).strip().lower())
    return variants


def _row_is_active(value: Any) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in YES_VALUES


def get_source_rows(workbook_data: Dict[str, List[Dict[str, Any]]], client_row: Dict[str, Any]) -> List[Dict[str, Any]]:
    variants = _client_key_variants(client_row)
    rows = []
    for row in workbook_data["sources"]:
        key = str(row.get("client_id_or_name") or "").strip().lower()
        if key not in variants:
            continue
        if not _row_is_active(row.get("active")):
            continue
        source_class = str(row.get("source_class") or "").strip()
        if source_class == "continuity":
            row = dict(row)
            row["source_class"] = "narrative_memory"
        if row["source_class"] not in SUPPORTED_SOURCE_CLASSES:
            raise ValueError(f"Unsupported source_class: {row['source_class']}")
        rows.append(row)
    return rows


def get_producer_note_rows(workbook_data: Dict[str, List[Dict[str, Any]]], client_row: Dict[str, Any]) -> List[Dict[str, Any]]:
    variants = _client_key_variants(client_row)
    rows = []
    for row in workbook_data["producer_notes"]:
        key = str(row.get("client_id_or_name") or "").strip().lower()
        if key not in variants:
            continue
        if not _row_is_active(row.get("active")):
            continue
        if not row.get("note_text"):
            continue
        rows.append(row)
    return rows


def build_client_memory_stub(client_row: Dict[str, Any], producer_notes: List[Dict[str, Any]]) -> Dict[str, Any]:
    client_id = str(client_row.get("client_id_optional") or client_row["client_name"]).strip().lower().replace(" ", "-")
    caveats = []
    notes = client_row.get("registry_notes_optional")
    if notes:
        caveats.append(str(notes))
    sensitivity_markers = []
    for note in producer_notes:
        if str(note.get("note_type") or "").strip() == "ClientNote":
            sensitivity_markers.append(str(note.get("title") or note.get("note_text") or "").strip())
    return {
        "client_id": client_id,
        "client_name": client_row["client_name"],
        "industry": "unknown",
        "service_mix": [],
        "north_star_goals": [],
        "preferred_language_style": "default_dp_profile",
        "what_counts_as_a_win": [],
        "recurring_concerns": [],
        "recurring_blockers": [],
        "sensitive_topics": sensitivity_markers,
        "approval_patterns": [],
        "known_reporting_caveats": caveats,
        "preferred_workstream_order": [],
        "sentiment_markers": [],
        "last_updated": date.today().isoformat(),
    }


def _make_external_source(ref_id: str, url: str, source_class: str, source_type: str, label: Optional[str] = None, extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    descriptor = {
        "kind": "external_ref",
        "location": url,
        "source_class": source_class,
        "source_type": source_type,
    }
    if label:
        descriptor["label"] = label
    if extra:
        descriptor.update(extra)
    return descriptor


def build_stage2_registry_payload(client_row: Dict[str, Any], source_rows: List[Dict[str, Any]], producer_notes: List[Dict[str, Any]]) -> Dict[str, Any]:
    client_id = str(client_row.get("client_id_optional") or client_row["client_name"]).strip().lower().replace(" ", "-")
    sources: Dict[str, Dict[str, Any]] = {}

    def add_source(ref_id: str, descriptor: Dict[str, Any]) -> str:
        sources[ref_id] = descriptor
        return ref_id

    alignment_ref = add_source(
        f"{client_id}:alignment_notes_doc",
        _make_external_source(
            f"{client_id}:alignment_notes_doc",
            client_row["alignment_notes_doc_link"],
            "narrative_memory",
            "alignment_notes_doc",
            label="Alignment Notes Master Doc",
        ),
    )
    kickoff_ref = add_source(f"{client_id}:kickoff", _make_external_source(f"{client_id}:kickoff", client_row["kickoff_doc_link"], "foundational_strategy", "kickoff_doc", label="Kickoff Doc"))
    acct_info_ref = add_source(f"{client_id}:acct_info", _make_external_source(f"{client_id}:acct_info", client_row["acct_info_doc_link"], "foundational_strategy", "acct_info_doc", label="Account Info Doc"))
    dossier_ref = add_source(f"{client_id}:dossier", _make_external_source(f"{client_id}:dossier", client_row["client_dosssier_doc_link"], "foundational_strategy", "client_dossier_doc", label="Client Dossier Doc"))
    overview_ref = add_source(f"{client_id}:overview", _make_external_source(f"{client_id}:overview", client_row["client_overview_link"], "foundational_strategy", "client_overview_doc", label="Client Overview Doc"))
    irg_ref = add_source(f"{client_id}:irg", _make_external_source(f"{client_id}:irg", client_row["irg_doc_link"], "foundational_strategy", "irg_doc", label="IRG Doc"))
    initial_strategy_ref = add_source(f"{client_id}:initial_strategy", _make_external_source(f"{client_id}:initial_strategy", client_row["initial_strategy_link"], "foundational_strategy", "initial_strategy_doc", label="Initial Strategy"))
    current_quarter_ref = add_source(f"{client_id}:quarterly_strategy", _make_external_source(f"{client_id}:quarterly_strategy", client_row["current_quarterly_strategy_link"], "active_strategy", "current_quarterly_strategy", label="Current Quarterly Strategy"))
    basecamp_internal_ref = add_source(f"{client_id}:basecamp_internal", _make_external_source(f"{client_id}:basecamp_internal", str(client_row["basecamp_internal_ref"]), "live_execution_truth", "basecamp_internal", label="Basecamp Internal"))
    basecamp_joint_ref = add_source(f"{client_id}:basecamp_joint", _make_external_source(f"{client_id}:basecamp_joint", str(client_row["basecamp_joint_ref"]), "live_execution_truth", "basecamp_joint", label="Basecamp Joint"))
    ga4_ref = add_source(f"{client_id}:ga4", _make_external_source(f"{client_id}:ga4", str(client_row["ga4_target_ref"]), "live_execution_truth", "ga4_target", label="GA4 Target"))
    gsc_ref = add_source(f"{client_id}:gsc", _make_external_source(f"{client_id}:gsc", str(client_row["gsc_target_ref"]), "live_execution_truth", "gsc_target", label="GSC Target"))

    reporting_refs = [ga4_ref, gsc_ref]
    if client_row.get("primary_dashboard_ref_optional"):
        reporting_refs.append(
            add_source(
                f"{client_id}:primary_dashboard",
                _make_external_source(f"{client_id}:primary_dashboard", client_row["primary_dashboard_ref_optional"], "live_execution_truth", "primary_dashboard", label="Primary Dashboard"),
            )
        )

    prior_report_refs: List[str] = []
    extra_note_refs: List[str] = []
    extra_reporting_refs: List[str] = []
    extra_strategy_refs: List[str] = []

    for idx, row in enumerate(source_rows, start=1):
        source_type = row["source_type"]
        source_class = row["source_class"]
        ref_id = f"{client_id}:source:{idx}"
        descriptor = _make_external_source(
            ref_id,
            row["source_ref"],
            source_class,
            source_type,
            label=row.get("source_label") or source_type,
            extra={
                "sub_ref_or_tab_optional": row.get("sub_ref_or_tab_optional"),
                "priority": row.get("priority"),
                "notes_optional": row.get("notes_optional"),
            },
        )
        add_source(ref_id, descriptor)
        if source_type == "prior_report":
            prior_report_refs.append(ref_id)
        elif source_type == "prior_call_note":
            extra_note_refs.append(ref_id)
        elif source_type == "dashboard":
            extra_reporting_refs.append(ref_id)
        elif source_class in {"foundational_strategy", "active_strategy"}:
            extra_strategy_refs.append(ref_id)

    reporting_refs.extend(extra_reporting_refs)

    producer_note_refs = []
    for idx, row in enumerate(producer_notes, start=1):
        ref_id = f"{client_id}:producer_note:{idx}"
        descriptor = {
            "kind": "inline",
            "content": row["note_text"],
            "source_class": "human_nuance",
            "source_type": row.get("note_type") or "producer_note",
            "label": row.get("title") or ref_id,
            "workstream_optional": row.get("workstream_optional"),
            "priority": row.get("priority"),
            "pinned": row.get("pinned"),
            "author": row.get("author"),
        }
        add_source(ref_id, descriptor)
        producer_note_refs.append(ref_id)

    memory_stub = build_client_memory_stub(client_row, producer_notes)
    memory_ref = f"{client_id}:client_memory_stub"
    sources[memory_ref] = {
        "kind": "inline",
        "content": json.dumps(memory_stub),
        "source_class": "human_nuance",
        "source_type": "client_memory_stub",
        "label": "Client Memory Stub",
    }

    client_record = {
        "client_id": client_id,
        "client_name": client_row["client_name"],
        "strategy_doc_refs": [kickoff_ref, acct_info_ref, dossier_ref, overview_ref, irg_ref, initial_strategy_ref] + extra_strategy_refs,
        "quarterly_strategy_ref": current_quarter_ref,
        "overview_doc_refs": [overview_ref, acct_info_ref, dossier_ref, irg_ref],
        "prior_report_refs": prior_report_refs,
        "prior_call_note_refs": [alignment_ref] + extra_note_refs,
        "basecamp_refs": [basecamp_internal_ref, basecamp_joint_ref],
        "reporting_refs": reporting_refs,
        "producer_note_refs": producer_note_refs,
        "template_ref": alignment_ref,
        "memory_ref": memory_ref,
    }
    return {
        "clients": {
            client_id: client_record,
        },
        "sources": sources,
    }


def build_live_run_payload(workbook_path: str | Path, client_selector: str) -> Dict[str, Any]:
    workbook_data = load_workbook_data(workbook_path)
    client_row = find_client(workbook_data, client_selector)
    source_rows = get_source_rows(workbook_data, client_row)
    producer_notes = get_producer_note_rows(workbook_data, client_row)
    registry_payload = build_stage2_registry_payload(client_row, source_rows, producer_notes)
    client_id = next(iter(registry_payload["clients"]))
    return {
        "client_id": client_id,
        "client_row": client_row,
        "source_rows": source_rows,
        "producer_notes": producer_notes,
        "stage2_registry_payload": registry_payload,
        "stage2_client_record": registry_payload["clients"][client_id],
    }


def write_live_run_artifacts(payload: Dict[str, Any], output_dir: str | Path) -> Dict[str, str]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    client_id = payload["client_id"]
    client_dir = output_path / client_id
    client_dir.mkdir(parents=True, exist_ok=True)

    files = {
        "client_row": client_dir / "client_row.json",
        "source_rows": client_dir / "source_rows.json",
        "producer_notes": client_dir / "producer_notes.json",
        "stage2_registry": client_dir / "stage2_registry_live.json",
        "stage2_client_record": client_dir / "stage2_client_record.json",
    }

    files["client_row"].write_text(json.dumps(payload["client_row"], indent=2))
    files["source_rows"].write_text(json.dumps(payload["source_rows"], indent=2))
    files["producer_notes"].write_text(json.dumps(payload["producer_notes"], indent=2))
    files["stage2_registry"].write_text(json.dumps(payload["stage2_registry_payload"], indent=2))
    files["stage2_client_record"].write_text(json.dumps(payload["stage2_client_record"], indent=2))

    return {key: str(path) for key, path in files.items()}
