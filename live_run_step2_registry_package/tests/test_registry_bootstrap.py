
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

import openpyxl

PACKAGE_ROOT = Path(__file__).resolve().parents[1]
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

from integrations.registry_adapter import build_live_run_payload, write_live_run_artifacts


def _build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Client Registry"
    ws.append([
        "client_id_optional","client_name","dp_owner","next_call_date","alignment_notes_doc_link","kickoff_doc_link",
        "acct_info_doc_link","client_dosssier_doc_link","client_overview_link","irg_doc_link","initial_strategy_link",
        "current_quarterly_strategy_link","basecamp_internal_ref","basecamp_joint_ref","ga4_target_ref","gsc_target_ref",
        "primary_dashboard_ref_optional","producer_notes_store_key_optional","last_verified_date","registry_notes_optional"
    ])
    ws.append([
        "client-1","Client One","DP",None,"https://docs/notes","https://docs/kickoff","https://docs/acct","https://docs/dossier",
        "https://docs/overview","https://docs/irg","https://docs/initial","https://docs/quarter","https://basecamp/internal",
        "https://basecamp/joint","12345","https://gsc/property","https://dash",None,None,"Caveat"
    ])

    ws2 = wb.create_sheet("Source Index")
    ws2.append(["client_id_or_name","source_type","source_class","source_label","source_ref","sub_ref_or_tab_optional","priority","active","notes_optional"])
    ws2.append(["client-1","prior_report","continuity","Feb Report","https://report/feb",None,"important","yes",None])

    ws3 = wb.create_sheet("Producer Notes")
    ws3.append(["note_id","client_id_or_name","note_type","title","note_text","workstream_optional","related_source_ref_optional","priority","pinned","starts_on_optional","expires_on_optional","author","active","notes_optional"])
    ws3.append(["note-1","client-1","RunNote","Lead with approvals","Lead with approvals first",None,None,"high","yes",None,None,"DP","yes",None])

    wb.save(path)


def main() -> None:
    with tempfile.TemporaryDirectory() as td:
        workbook_path = Path(td) / "workbook.xlsx"
        _build_workbook(workbook_path)
        payload = build_live_run_payload(workbook_path, "client-1")
        assert payload["client_id"] == "client-1"
        assert payload["source_rows"][0]["source_class"] == "narrative_memory"
        output = write_live_run_artifacts(payload, Path(td) / "out")
        registry_path = Path(output["stage2_registry"])
        data = json.loads(registry_path.read_text())
        assert "client-1" in data["clients"]
        assert data["clients"]["client-1"]["prior_report_refs"]
        assert data["clients"]["client-1"]["producer_note_refs"]
    print("Step 2 registry smoke test passed.")


if __name__ == "__main__":
    main()
